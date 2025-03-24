from openpyxl import Workbook, load_workbook
import os


#os.path.dirname(__file__)
#os.chdir()
def create_Excel():
    name = "data.xlsx"

    wb = Workbook()

    sheet = wb.active

    sheet["A1"] = "Username"
    sheet["B1"] = "Password"
    sheet["A2"] = "Paul"
    sheet["B2"] = 12345

    wb.save(filename=name)


def read_data():
    wb = load_workbook()
    data = "Paul, 12345"
    return data


if __name__ == "__main__":
    create_Excel()